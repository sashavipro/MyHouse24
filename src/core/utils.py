"""src/core/utils.py."""

from contextlib import suppress

import openpyxl
from django.utils import timezone

from src.finance.models import PaymentDetails


class ReceiptExcelGenerator:
    """Receipt Excel Generator."""

    def __init__(self, receipt, template_path):
        """Init."""
        self.receipt = receipt
        self.template_path = template_path

    def generate_workbook(self):
        """Generate the Excel workbook."""  # FIXED: D401 - Imperative mood
        workbook = openpyxl.load_workbook(self.template_path)
        sheet = workbook.active

        context = self._build_template_context()
        self._replace_template_variables(sheet, context)

        items = self.receipt.receiptitem_set.select_related("service__unit").all()
        template_row_idx = self._find_template_row(sheet)

        if template_row_idx and items:
            self._insert_receipt_items(sheet, template_row_idx, items)

        return workbook

    def _build_template_context(self):
        """Build context dictionary for template variable replacement."""
        receipt = self.receipt
        status_map = {
            "paid": "Оплачена",
            "partially_paid": "Частично оплачена",
            "unpaid": "Неоплачена",
        }

        period = "—"
        if receipt.period_start and receipt.period_end:
            period = (
                f"{receipt.period_start.strftime('%d.%m.%Y')} "
                f"- {receipt.period_end.strftime('%d.%m.%Y')}"
            )

        try:
            payment_details = PaymentDetails.objects.get(pk=1)
            company_name = payment_details.company_name
            company_details = payment_details.info
        except PaymentDetails.DoesNotExist:
            company_name = "Название компании не задано"
            company_details = "Реквизиты не заданы"

        account_balance = 0.00
        if receipt.apartment and receipt.apartment.personal_account:
            account_balance = receipt.apartment.personal_account.balance

        total_to_pay = receipt.total_amount
        if account_balance < 0:
            total_to_pay += abs(account_balance)

        return {
            "{{ receipt.is_posted }}": (
                "Conducted" if receipt.is_posted else "Not conducted"
            ),
            "{{ receipt.status }}": status_map.get(receipt.status, "—"),
            "{{ receipt.period }}": period,
            "{{ personal_account.number }}": (
                receipt.apartment.personal_account.number
                if receipt.apartment and receipt.apartment.personal_account
                else "—"
            ),
            "{{ owner.phone }}": (
                receipt.apartment.owner.phone
                if receipt.apartment.owner and receipt.apartment.owner.phone
                else "—"
            ),
            "{{ apartment.house }}": receipt.apartment.house.title,
            "{{ house.address }}": receipt.apartment.house.address,
            "{{ apartment.number }}": receipt.apartment.number,
            "{{ apartment.section }}": (
                receipt.apartment.section.name if receipt.apartment.section else "—"
            ),
            "{{ receipt.tariff }}": receipt.tariff.name if receipt.tariff else "—",
            "{{ receipt.total_amount }}": f"{receipt.total_amount:.2f}",
            "{{ receipt.number }}": receipt.number,
            "{{ receipt.date }}": receipt.date.strftime("%d.%m.%Y"),
            "{{ owner.full_name }}": (
                receipt.apartment.owner.get_full_name()
                if receipt.apartment.owner
                else "—"
            ),
            "{{ pay_company.name }}": company_name,
            "{{ pay_company.details }}": company_details,
            "{{ account.balance }}": f"{account_balance:.2f}",
            "{{ total_pay }}": f"{total_to_pay:.2f}",
            "{{ receipt.created_date }}": timezone.now().strftime("%d.%m.%Y"),
        }

    def _replace_template_variables(self, sheet, context):
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for key, value in context.items():
                        if key in cell.value:
                            cell.value = cell.value.replace(key, str(value))

    def _find_template_row(self, sheet):
        for i, row in enumerate(sheet.iter_rows(), 1):
            for cell in row:
                if isinstance(cell.value, str) and (
                    "{{ item.name }}" in cell.value or "{{ item.number }}" in cell.value
                ):
                    return i
        return None

    def _insert_receipt_items(self, sheet, template_row_idx, items):
        template_row = sheet[template_row_idx]

        # FIXED: PERF401 - Use list comprehension instead of loop with append
        merged_ranges_to_copy = [
            {
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col,
            }
            for merged_range in sheet.merged_cells.ranges
            if merged_range.min_row <= template_row_idx <= merged_range.max_row
        ]

        for i, item in enumerate(items, 1):
            new_row_idx = template_row_idx + i
            sheet.insert_rows(new_row_idx)
            new_row = sheet[new_row_idx]

            self._copy_row_style(template_row, new_row)
            self._populate_item_row(new_row, template_row, item, i)

            # FIXED: SIM105 - Use contextlib.suppress instead of try-except-pass
            for merge_info in merged_ranges_to_copy:
                with suppress(ValueError):
                    sheet.merge_cells(
                        start_row=new_row_idx,
                        start_column=merge_info["min_col"],
                        end_row=new_row_idx,
                        end_column=merge_info["max_col"],
                    )

        sheet.delete_rows(template_row_idx)

    def _copy_row_style(self, template_row, new_row):
        for j, template_cell in enumerate(template_row, 1):
            new_cell = new_row[j - 1]
            if template_cell.has_style:
                new_cell.font = template_cell.font.copy()
                new_cell.border = template_cell.border.copy()
                new_cell.fill = template_cell.fill.copy()
                new_cell.number_format = template_cell.number_format
                new_cell.protection = template_cell.protection.copy()
                new_cell.alignment = template_cell.alignment.copy()

    def _populate_item_row(self, new_row, template_row, item, item_number):
        for j, template_cell in enumerate(template_row, 1):
            new_cell = new_row[j - 1]
            if not isinstance(template_cell.value, str):
                continue

            val_str = self._replace_item_placeholders(
                template_cell.value, item, item_number
            )
            try:
                numeric_placeholders = [
                    "{{ item.consumption }}",
                    "{{ item.price }}",
                    "{{ item.amount }}",
                    "{{ item.number }}",
                ]
                if any(p in template_cell.value for p in numeric_placeholders):
                    new_cell.value = float(val_str.replace(",", "."))
                else:
                    new_cell.value = val_str
            except (ValueError, TypeError):
                new_cell.value = val_str

    def _replace_item_placeholders(self, template_value, item, item_number):
        return (
            template_value.replace("{{ item.number }}", str(item_number))
            .replace("{{ item.name }}", item.service.name)
            .replace(
                "{{ item.consumption }}",
                f"{item.consumption:.3f}" if item.consumption is not None else "",
            )
            .replace("{{ item.unit }}", item.service.unit.name)
            .replace(
                "{{ item.price }}",
                f"{item.price_per_unit:.2f}" if item.price_per_unit is not None else "",
            )
            .replace(
                "{{ item.amount }}",
                f"{item.amount:.2f}" if item.amount is not None else "",
            )
        )
