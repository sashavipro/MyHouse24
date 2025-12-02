"""src/finance/views.py."""

import datetime
import io
import logging

import openpyxl
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.db import DatabaseError
from django.db import transaction
from django.db.models import IntegerField
from django.db.models import Max
from django.db.models import ProtectedError
from django.db.models import Q
from django.db.models import Sum
from django.db.models.functions import Cast
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView
from django.views.generic import DetailView
from django.views.generic import ListView
from django.views.generic import TemplateView
from django.views.generic import UpdateView
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.building.models import Apartment
from src.building.models import House
from src.building.models import PersonalAccount
from src.core.utils import ReceiptExcelGenerator
from src.finance.forms import ArticleForm
from src.finance.forms import CashBoxExpenseForm
from src.finance.forms import CashBoxIncomeForm
from src.finance.forms import CounterReadingForm
from src.finance.forms import PaymentDetailsForm
from src.finance.forms import PrintTemplateForm
from src.finance.forms import ReceiptForm
from src.finance.forms import ReceiptItemFormSet
from src.finance.forms import ServiceFormSet
from src.finance.forms import TariffForm
from src.finance.forms import TariffServiceFormSet
from src.finance.forms import UnitFormSet
from src.finance.models import Article
from src.finance.models import CashBox
from src.finance.models import Counter
from src.finance.models import CounterReading
from src.finance.models import PaymentDetails
from src.finance.models import PrintTemplate
from src.finance.models import Receipt
from src.finance.models import Service
from src.finance.models import Tariff
from src.finance.models import Unit
from src.users.models import Ticket
from src.users.models import User
from src.users.permissions import RoleRequiredMixin

from .tasks import send_receipt_email_task

logger = logging.getLogger(__name__)


class AdminStatsView(LoginRequiredMixin, RoleRequiredMixin, TemplateView):
    """Display the main admin panel dashboard."""

    template_name = "core/adminlte/admin_stats.html"
    permission_required = "has_statistics"

    def get_context_data(self, **kwargs):
        """Get context data."""
        context = super().get_context_data(**kwargs)

        context["houses_count"] = House.objects.count()
        context["active_owners_count"] = User.objects.filter(
            user_type=User.UserType.OWNER, status=User.UserStatus.ACTIVE
        ).count()
        context["tickets_in_progress_count"] = Ticket.objects.filter(
            status=Ticket.TicketStatus.IN_PROGRESS
        ).count()
        context["apartments_count"] = Apartment.objects.count()
        context["personal_accounts_count"] = PersonalAccount.objects.count()
        context["tickets_new_count"] = Ticket.objects.filter(
            status=Ticket.TicketStatus.NEW
        ).count()

        context["total_debt_accounts"] = (
            PersonalAccount.objects.filter(balance__lt=0).aggregate(s=Sum("balance"))[
                "s"
            ]
            or 0
        )
        context["total_balance_accounts"] = (
            PersonalAccount.objects.aggregate(s=Sum("balance"))["s"] or 0
        )

        income_total = (
            CashBox.objects.filter(
                is_posted=True, article__type=Article.ArticleType.INCOME
            ).aggregate(s=Sum("amount"))["s"]
            or 0
        )
        expense_total = (
            CashBox.objects.filter(
                is_posted=True, article__type=Article.ArticleType.EXPENSE
            ).aggregate(s=Sum("amount"))["s"]
            or 0
        )
        context["cashbox_balance"] = income_total - expense_total

        today = timezone.now()
        current_year = today.year

        months_labels = []
        for month in range(1, 13):
            date_obj = datetime.date(current_year, month, 1)
            months_labels.append(date_obj.strftime("%b., %Y"))

        context["chart_labels"] = months_labels

        receipts_data = (
            Receipt.objects.filter(date__year=current_year, is_posted=True)
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(total=Sum("total_amount"))
            .order_by("month")
        )

        payments_data = (
            CashBox.objects.filter(
                date__year=current_year,
                is_posted=True,
                article__type=Article.ArticleType.INCOME,
            )
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(total=Sum("amount"))
            .order_by("month")
        )

        context["chart_receipts_debt"] = self._fill_monthly_data(receipts_data)
        context["chart_receipts_paid"] = self._fill_monthly_data(payments_data)

        cashbox_income_data = payments_data

        cashbox_expense_data = (
            CashBox.objects.filter(
                date__year=current_year,
                is_posted=True,
                article__type=Article.ArticleType.EXPENSE,
            )
            .annotate(month=TruncMonth("date"))
            .values("month")
            .annotate(total=Sum("amount"))
            .order_by("month")
        )

        context["chart_cashbox_income"] = self._fill_monthly_data(cashbox_income_data)
        context["chart_cashbox_expense"] = self._fill_monthly_data(cashbox_expense_data)

        return context

    def _fill_monthly_data(self, queryset_data):
        """Map DB aggregation result to a list of 12 months."""
        data_map = {
            entry["month"].strftime("%m"): float(entry["total"])
            for entry in queryset_data
            if entry["month"]
        }

        result = []
        for i in range(1, 13):
            key = f"{i:02d}"
            result.append(data_map.get(key, 0.0))
        return result


class ManageServicesView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Processes GET and POST requests for managing Services and Units of Measure.

    With correct handling of deletion errors.
    """

    template_name = "core/adminlte/admin_services.html"
    permission_required = "has_service"

    def get(self, request, *args, **kwargs):
        """Display form sets for editing."""
        service_formset = ServiceFormSet(
            queryset=Service.objects.all().order_by("name"), prefix="services"
        )
        unit_formset = UnitFormSet(
            queryset=Unit.objects.all().order_by("name"), prefix="units"
        )

        context = {
            "service_formset": service_formset,
            "unit_formset": unit_formset,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Save changes for one of the form sets."""
        if "save_services" in request.POST:
            formset = ServiceFormSet(request.POST, prefix="services")
            if formset.is_valid():
                try:
                    with transaction.atomic():
                        formset.save()
                    return redirect("finance:manage_services")
                except ProtectedError:
                    messages.error(
                        request,
                        "Невозможно удалить услугу, так как она используется в "
                        "квитанциях, счетчиках или тарифах.",
                    )

            unit_formset = UnitFormSet(
                queryset=Unit.objects.all().order_by("name"), prefix="units"
            )
            return render(
                request,
                self.template_name,
                {"service_formset": formset, "unit_formset": unit_formset},
            )

        if "save_units" in request.POST:
            formset = UnitFormSet(request.POST, prefix="units")
            if formset.is_valid():
                try:
                    with transaction.atomic():
                        formset.save()
                    return redirect("finance:manage_services")
                except ProtectedError:
                    messages.error(
                        request,
                        "It is not possible to delete the unit of measurement "
                        "because it is used in one or more services.",
                    )

            service_formset = ServiceFormSet(
                queryset=Service.objects.all().order_by("name"), prefix="services"
            )
            return render(
                request,
                self.template_name,
                {"service_formset": service_formset, "unit_formset": formset},
            )

        return redirect("finance:manage_services")


class TariffListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Display a list of tariffs."""

    model = Tariff
    template_name = "core/adminlte/tariff_list.html"
    permission_required = "has_tariff"


class TariffDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Display detailed information about a single tariff."""

    model = Tariff
    template_name = "core/adminlte/tariff_detail.html"
    context_object_name = "tariff"
    permission_required = "has_tariff"


class TariffCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    """Handle the creation of a new tariff."""

    model = Tariff
    form_class = TariffForm
    template_name = "core/adminlte/tariff_form.html"
    permission_required = "has_tariff"

    def get_success_url(self):
        """Redirect to the detail page of the newly created tariff."""
        return reverse_lazy("finance:tariff_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        """Add the services formset to the context."""
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["services_formset"] = TariffServiceFormSet(
                self.request.POST, prefix="services"
            )
        else:
            context["services_formset"] = TariffServiceFormSet(prefix="services")
        return context

    def form_valid(self, form):
        """Save the main form and the related formset if both are valid."""
        context = self.get_context_data(form=form)
        services_formset = context["services_formset"]
        if services_formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                services_formset.instance = self.object
                services_formset.save()
            return redirect(self.get_success_url())
        return self.form_invalid(form)


class TariffUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    """Handle updating an existing tariff."""

    model = Tariff
    form_class = TariffForm
    template_name = "core/adminlte/tariff_form.html"
    permission_required = "has_tariff"

    def get_success_url(self):
        """Redirect to the detail page of the updated tariff."""
        return reverse_lazy("finance:tariff_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        """Add the pre-filled services formset to the context."""
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["services_formset"] = TariffServiceFormSet(
                self.request.POST, instance=self.object, prefix="services"
            )
        else:
            context["services_formset"] = TariffServiceFormSet(
                instance=self.object, prefix="services"
            )
        return context

    def form_valid(self, form):
        """Save the main form and the related formset if both are valid."""
        context = self.get_context_data(form=form)
        services_formset = context["services_formset"]
        if services_formset.is_valid():
            with transaction.atomic():
                self.object = form.save()
                services_formset.instance = self.object
                services_formset.save()
            return redirect(self.get_success_url())
        return self.form_invalid(form)


class ArticleListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Display the page with the list of payment articles."""

    model = Article
    template_name = "core/adminlte/article_list.html"
    permission_required = "has_article"


class ArticleCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    """Display the form for creating a new article."""

    model = Article
    form_class = ArticleForm
    template_name = "core/adminlte/article_form.html"
    success_url = reverse_lazy("finance:article_list")
    permission_required = "has_article"


class ArticleUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    """Display the form for editing an existing article."""

    model = Article
    form_class = ArticleForm
    template_name = "core/adminlte/article_form.html"
    success_url = reverse_lazy("finance:article_list")
    permission_required = "has_article"


class PaymentDetailsUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    """Handle displaying and updating the singleton PaymentDetails object."""

    model = PaymentDetails
    form_class = PaymentDetailsForm
    template_name = "core/adminlte/payment_details_form.html"
    success_url = reverse_lazy("finance:payment_details")
    permission_required = "has_payment_details"

    def get_object(self, queryset=None):
        """Return the single PaymentDetails instance, creating it if needed.

        This implements the singleton pattern for PaymentDetails.
        """
        obj, _created = self.model.objects.get_or_create(pk=1)
        return obj


class CounterListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Display a page with a list of all counters."""

    model = Counter
    template_name = "core/adminlte/counter_list.html"
    permission_required = "has_counters"

    def get_context_data(self, **kwargs):
        """Add data for filters to the context."""
        context = super().get_context_data(**kwargs)
        context["houses"] = House.objects.all().order_by("title")
        context["services"] = Service.objects.filter(show_in_counters=True).order_by(
            "name"
        )
        return context


class CounterReadingListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Show a GENERAL list of all indications with the option to filter."""

    model = CounterReading
    template_name = "core/adminlte/counter_reading_list.html"
    permission_required = "has_counters"

    def get_context_data(self, **kwargs):
        """Add data for filter dropdown lists to the context."""
        context = super().get_context_data(**kwargs)
        context["houses"] = House.objects.all().order_by("title")
        context["services"] = Service.objects.filter(show_in_counters=True).order_by(
            "name"
        )
        context["statuses"] = CounterReading.CounterStatus.choices

        return context


class CounterReadingCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    """Processe the creation of a new meter reading."""

    model = CounterReading
    form_class = CounterReadingForm
    template_name = "core/adminlte/counter_reading_form.html"
    permission_required = "has_counters"

    def get_success_url(self):
        """Define the URL for redirection after successful saving."""
        if "save_and_add_new" in self.request.POST:
            try:
                current_apartment = self.object.apartment
                next_apartment = (
                    Apartment.objects.filter(
                        house=current_apartment.house,
                        number__gt=current_apartment.number,
                    )
                    .order_by("number")
                    .first()
                )

                if next_apartment:
                    return (
                        reverse_lazy("finance:counter_reading_add")
                        + f"?apartment={next_apartment.pk}"
                    )
            except (AttributeError, Apartment.DoesNotExist):
                pass
            return reverse_lazy("finance:counter_reading_add")

        return (
            reverse_lazy("finance:counter_reading_list")
            + f"?counter={self.object.counter.pk}"
        )

    def get_context_data(self, **kwargs):
        """Add to the context of the home for the drop-down list."""
        context = super().get_context_data(**kwargs)
        context["houses"] = House.objects.all().order_by("title")
        if hasattr(self, "initial_data"):
            context["initial_data"] = self.initial_data
        return context

    def get_initial(self):
        """Prepare data for pre-filling fields in the form."""
        initial = super().get_initial()
        initial["date"] = timezone.localdate()

        result = CounterReading.objects.aggregate(
            max_number=Max(Cast("number", output_field=IntegerField()))
        )
        last_number = result.get("max_number") or 0

        initial["number"] = str(last_number + 1)

        apartment_id = self.request.GET.get("apartment")
        if apartment_id:
            try:
                apartment = Apartment.objects.select_related("house", "section").get(
                    pk=apartment_id
                )

                self.initial_data = {
                    "house_id": apartment.house_id,
                    "section_id": apartment.section_id if apartment.section else None,
                    "apartment_id": apartment.pk,
                }
            except Apartment.DoesNotExist:
                pass
        return initial

    def form_valid(self, form):
        """Find or creates a counter before saving the reading."""
        apartment = form.cleaned_data.get("apartment")
        service = form.cleaned_data.get("service")

        if not apartment or not service:
            form.add_error(None, "Необходимо выбрать квартиру и услугу.")
            return self.form_invalid(form)

        try:
            with transaction.atomic():
                counter, _created = Counter.objects.get_or_create(
                    apartment=apartment,
                    service=service,
                    defaults={"serial_number": f"auto-{apartment.pk}-{service.pk}"},
                )

                reading = form.save(commit=False)
                reading.counter = counter

                self.object = reading
                reading.save()
        except (DatabaseError, ValidationError):
            logger.exception("Error saving counter reading")
            form.add_error(None, "Ошибка при сохранении показаний.")
            return self.form_invalid(form)

        return redirect(self.get_success_url())


class CounterReadingUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    """Processe editing of existing meter readings."""

    model = CounterReading
    form_class = CounterReadingForm
    template_name = "core/adminlte/counter_reading_form.html"
    permission_required = "has_counters"

    def get_success_url(self):
        """After editing, always return to the history page."""
        return (
            reverse_lazy("finance:counter_reading_list")
            + f"?counter={self.object.counter.pk}"
        )

    def get_context_data(self, **kwargs):
        """Add data for pre-filling fields in JS."""
        context = super().get_context_data(**kwargs)
        context["houses"] = House.objects.all().order_by("title")
        context["initial_data"] = {
            "house_id": self.object.apartment.house_id,
            "section_id": self.object.apartment.section_id
            if self.object.apartment.section
            else None,
            "apartment_id": self.object.apartment.pk,
        }
        return context

    def form_valid(self, form):
        """Find or creates a counter before saving the reading."""
        apartment = form.cleaned_data.get("apartment")
        service = form.cleaned_data.get("service")

        if not apartment or not service:
            form.add_error(None, "Необходимо выбрать квартиру и услугу.")
            return self.form_invalid(form)

        try:
            with transaction.atomic():
                counter, _created = Counter.objects.get_or_create(
                    apartment=apartment,
                    service=service,
                    defaults={"serial_number": f"auto-{apartment.pk}-{service.pk}"},
                )

                reading = form.save(commit=False)
                reading.counter = counter

                self.object = reading
                reading.save()
        except (DatabaseError, ValidationError):
            logger.exception("Error updating counter reading")
            form.add_error(None, "Ошибка при обновлении показаний.")
            return self.form_invalid(form)

        return redirect(self.get_success_url())


class ReceiptListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """Display a list of receipts with filters and statistics."""

    model = Receipt
    template_name = "core/adminlte/receipt_list.html"
    permission_required = "has_receipt"

    def get_context_data(self, **kwargs):
        """Add statistics and filter data to the context."""
        context = super().get_context_data(**kwargs)

        income = (
            CashBox.objects.filter(is_posted=True, article__type="income").aggregate(
                s=Sum("amount")
            )["s"]
            or 0
        )
        expense = (
            CashBox.objects.filter(is_posted=True, article__type="expense").aggregate(
                s=Sum("amount")
            )["s"]
            or 0
        )
        context["total_cash"] = income - expense

        context["total_balance_accounts"] = (
            Apartment.objects.aggregate(total=Sum("personal_account__balance"))["total"]
            or 0
        )
        context["total_debt"] = (
            Apartment.objects.filter(personal_account__balance__lt=0).aggregate(
                total=Sum("personal_account__balance")
            )["total"]
            or 0
        )
        context["owners"] = User.objects.filter(user_type="owner")
        context["statuses"] = Receipt.ReceiptStatus.choices
        return context


class ReceiptDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Show detailed information for a single receipt."""

    model = Receipt
    template_name = "core/adminlte/receipt_detail.html"
    permission_required = "has_receipt"


class ReceiptCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    """Creating a new receipt."""

    model = Receipt
    form_class = ReceiptForm
    template_name = "core/adminlte/receipt_form.html"
    permission_required = "has_receipt"

    def get_success_url(self):
        """Redirect to the detail page of the newly created receipt."""
        return reverse_lazy("finance:receipt_detail", kwargs={"pk": self.object.pk})

    def get_initial(self):
        """Prepare initial data for the form.

        Including the next receipt number and apartment details from the GET
        parameter.
        """
        initial = super().get_initial()

        result = Receipt.objects.aggregate(
            max_number=Max(Cast("number", output_field=IntegerField()))
        )

        last_number = result.get("max_number") or 0

        next_number = last_number + 1

        initial["number"] = str(next_number)

        apartment_id = self.request.GET.get("apartment")
        if apartment_id:
            try:
                apartment = Apartment.objects.select_related(
                    "house", "section", "personal_account", "tariff"
                ).get(pk=apartment_id)

                initial["apartment"] = apartment
                if apartment.tariff:
                    initial["tariff"] = apartment.tariff

                self.initial_data = {
                    "house_id": apartment.house_id,
                    "section_id": apartment.section_id if apartment.section else None,
                    "apartment_id": apartment.pk,
                }
            except Apartment.DoesNotExist:
                pass

        return initial

    def get_context_data(self, **kwargs):
        """Add data to the context for JavaScript pre-filling."""
        context = super().get_context_data(**kwargs)
        context["houses"] = House.objects.order_by("title")
        context["tariffs"] = Tariff.objects.order_by("name")

        if hasattr(self, "initial_data"):
            context["initial_data"] = self.initial_data

        if self.request.POST:
            context["services_formset"] = ReceiptItemFormSet(
                self.request.POST, prefix="services"
            )
        else:
            context["services_formset"] = ReceiptItemFormSet(prefix="services")
        return context

    def form_valid(self, form):
        """Save the receipt, calculate total, and update counter readings."""
        context = self.get_context_data(form=form)
        services_formset = context["services_formset"]

        if services_formset.is_valid():
            try:
                with transaction.atomic():
                    self.object = form.save(commit=False)
                    self.object.total_amount = 0
                    self.object.save()

                    services_formset.instance = self.object
                    services_formset.save()

                    total = (
                        self.object.receiptitem_set.aggregate(total=Sum("amount"))[
                            "total"
                        ]
                        or 0
                    )
                    self.object.total_amount = total
                    self.object.save()

                    service_ids = self.object.receiptitem_set.values_list(
                        "service_id", flat=True
                    )

                    if self.object.apartment:
                        readings_to_update = CounterReading.objects.filter(
                            counter__apartment=self.object.apartment,
                            counter__service_id__in=service_ids,
                            status=CounterReading.CounterStatus.NEW,
                        )
                        readings_to_update.update(
                            status=CounterReading.CounterStatus.CONSIDERED
                        )

            except (DatabaseError, ValidationError):
                logger.exception("Error creating receipt")
                return self.form_invalid(form)

            return redirect(self.get_success_url())
        return self.form_invalid(form)


class ReceiptUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    """Editing an existing receipt with pre-filled fields."""

    model = Receipt
    form_class = ReceiptForm
    template_name = "core/adminlte/receipt_form.html"
    permission_required = "has_receipt"

    def get_success_url(self):
        """Redirect to the detail page of the updated receipt."""
        return reverse_lazy("finance:receipt_detail", kwargs={"pk": self.object.pk})

    def get_context_data(self, **kwargs):
        """Add formset and JavaScript pre-fill data to the template context."""
        context = super().get_context_data(**kwargs)
        context["houses"] = House.objects.order_by("title")
        context["tariffs"] = Tariff.objects.order_by("name")

        if self.object.apartment:
            apartment = self.object.apartment
            context["initial_data"] = {
                "house_id": apartment.house_id,
                "section_id": apartment.section_id if apartment.section else None,
                "apartment_id": apartment.pk,
            }

        if self.request.POST:
            context["services_formset"] = ReceiptItemFormSet(
                self.request.POST, instance=self.object, prefix="services"
            )
        else:
            context["services_formset"] = ReceiptItemFormSet(
                instance=self.object, prefix="services"
            )
        return context

    def form_valid(self, form):
        """Save the receipt, calculate total, and update counter readings."""
        context = self.get_context_data(form=form)
        services_formset = context["services_formset"]

        if services_formset.is_valid():
            try:
                with transaction.atomic():
                    old_service_ids = set(
                        self.object.receiptitem_set.values_list("service_id", flat=True)
                    )

                    self.object = form.save(commit=False)
                    self.object.total_amount = 0
                    self.object.save()

                    services_formset.instance = self.object
                    services_formset.save()

                    new_service_ids = set(
                        self.object.receiptitem_set.values_list("service_id", flat=True)
                    )

                    total = (
                        self.object.receiptitem_set.aggregate(total=Sum("amount"))[
                            "total"
                        ]
                        or 0
                    )
                    self.object.total_amount = total
                    self.object.save()

                    if self.object.apartment:
                        added_services = new_service_ids - old_service_ids
                        if added_services:
                            CounterReading.objects.filter(
                                counter__apartment=self.object.apartment,
                                counter__service_id__in=added_services,
                                status=CounterReading.CounterStatus.NEW,
                            ).update(status=CounterReading.CounterStatus.CONSIDERED)

                        removed_services = old_service_ids - new_service_ids
                        if removed_services:
                            CounterReading.objects.filter(
                                counter__apartment=self.object.apartment,
                                counter__service_id__in=removed_services,
                                status=CounterReading.CounterStatus.CONSIDERED,
                            ).update(status=CounterReading.CounterStatus.NEW)

                        remained_services = new_service_ids & old_service_ids
                        if remained_services:
                            CounterReading.objects.filter(
                                counter__apartment=self.object.apartment,
                                counter__service_id__in=remained_services,
                                status=CounterReading.CounterStatus.NEW,
                            ).update(status=CounterReading.CounterStatus.CONSIDERED)

            except (DatabaseError, ValidationError):
                logger.exception("Error updating receipt")
                return self.form_invalid(form)

            return redirect(self.get_success_url())
        return self.form_invalid(form)


class ReceiptPrintFormView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Show template selection page and generate Excel file or send Email."""

    template_name = "core/adminlte/receipt_print_form.html"
    permission_required = "has_receipt"

    def get(self, request, *args, **kwargs):
        """Display the template selection form."""
        receipt = get_object_or_404(Receipt, pk=kwargs["pk"])
        templates = PrintTemplate.objects.all().order_by("-is_default", "name")
        return render(
            request, self.template_name, {"receipt": receipt, "templates": templates}
        )

    def post(self, request, *args, **kwargs):
        """Process the download or upload of the generated Excel file."""
        receipt = get_object_or_404(
            Receipt.objects.select_related(
                "apartment__house",
                "apartment__section",
                "apartment__owner",
                "apartment__personal_account",
                "tariff",
            ),
            pk=kwargs["pk"],
        )
        template_id = request.POST.get("template_id")

        if not template_id:
            messages.error(request, "Пожалуйста, выберите шаблон.")
            templates = PrintTemplate.objects.all().order_by("-is_default", "name")
            return render(
                request,
                self.template_name,
                {"receipt": receipt, "templates": templates},
            )

        if "send_email" in request.POST:
            if not receipt.apartment.owner or not receipt.apartment.owner.email:
                messages.error(
                    request,
                    "У владельца квартиры не указан Email. Отправка невозможна.",  # noqa: RUF001
                )
            else:
                send_receipt_email_task.delay(receipt.pk, int(template_id))

                messages.success(
                    request,
                    f"Запущен процесс отправки квитанции на "
                    f"{receipt.apartment.owner.email}."
                    f" Это может занять некоторое время.",
                )

            return redirect("finance:receipt_print_form", pk=receipt.pk)

        if "download" in request.POST:
            template_obj = get_object_or_404(PrintTemplate, pk=template_id)

            try:
                generator = ReceiptExcelGenerator(
                    receipt, template_obj.template_file.path
                )
                workbook = generator.generate_workbook()
                return self._create_download_response(workbook, receipt)

            except Exception:
                logger.exception("Error generating receipt Excel file")
                messages.error(request, "Ошибка при формировании файла квитанции.")

        templates = PrintTemplate.objects.all().order_by("-is_default", "name")
        return render(
            request, self.template_name, {"receipt": receipt, "templates": templates}
        )

    def _create_download_response(self, workbook, receipt):
        """Create HTTP response for downloading the Excel file."""
        virtual_workbook = io.BytesIO()
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)

        response = HttpResponse(
            virtual_workbook.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f'attachment; filename="receipt_{receipt.number}.xlsx"'
        )
        return response


class ReceiptTemplateSettingsView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Managing templates for printing receipts."""

    template_name = "core/adminlte/receipt_template_settings.html"
    permission_required = "has_receipt"

    def get(self, request, *args, **kwargs):
        """Handle GET requests for template management."""
        action = request.GET.get("action")
        template_id = request.GET.get("template_id")
        if template_id:
            template = get_object_or_404(PrintTemplate, pk=template_id)
            if action == "delete":
                template.delete()
                return redirect("finance:receipt_template_settings")
            if action == "set_default":
                template.is_default = True
                template.save()
                return redirect("finance:receipt_template_settings")

        form = PrintTemplateForm()
        templates = PrintTemplate.objects.all().order_by("-is_default", "name")
        return render(
            request, self.template_name, {"templates": templates, "form": form}
        )

    def post(self, request, *args, **kwargs):
        """Handle POST requests for creating new templates."""
        form = PrintTemplateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("finance:receipt_template_settings")

        templates = PrintTemplate.objects.all().order_by("-is_default", "name")
        return render(
            request, self.template_name, {"templates": templates, "form": form}
        )


class CashBoxListView(LoginRequiredMixin, RoleRequiredMixin, ListView):
    """CashBoxListView."""

    model = CashBox
    template_name = "core/adminlte/cashbox_list.html"
    permission_required = "has_cashbox"

    def get_context_data(self, **kwargs):
        """Get the context data for the template."""
        context = super().get_context_data(**kwargs)
        income = (
            CashBox.objects.filter(is_posted=True, article__type="income").aggregate(
                s=Sum("amount")
            )["s"]
            or 0
        )
        expense = (
            CashBox.objects.filter(is_posted=True, article__type="expense").aggregate(
                s=Sum("amount")
            )["s"]
            or 0
        )
        context["total_cash"] = income - expense

        context["total_balance_accounts"] = (
            Apartment.objects.aggregate(total=Sum("personal_account__balance"))["total"]
            or 0
        )
        context["total_debt"] = (
            Apartment.objects.filter(personal_account__balance__lt=0).aggregate(
                total=Sum("personal_account__balance")
            )["total"]
            or 0
        )
        context["articles"] = Article.objects.all()
        context["owners"] = User.objects.filter(user_type=User.UserType.OWNER)
        return context


class CashBoxIncomeCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    """CashBoxIncomeCreateView."""

    model = CashBox
    form_class = CashBoxIncomeForm
    template_name = "core/adminlte/cashbox_income_form.html"
    success_url = reverse_lazy("finance:cashbox_list")
    permission_required = "has_cashbox"

    def get_initial(self):
        """Get the initial data for the template."""
        initial = super().get_initial()
        last = CashBox.objects.aggregate(Max("id"))["id__max"] or 0
        initial["number"] = f"{last + 1}".zfill(10)
        initial["date"] = timezone.now().date()
        initial["manager"] = self.request.user

        pa_id = self.request.GET.get("personal_account_id")
        if pa_id:
            from src.building.models import PersonalAccount

            try:
                pa = PersonalAccount.objects.select_related("apartment__owner").get(
                    pk=pa_id
                )
                initial["personal_account"] = pa

                if pa.apartment and pa.apartment.owner:
                    initial["owner"] = pa.apartment.owner
            except PersonalAccount.DoesNotExist:
                pass

        return initial

    def form_valid(self, form):
        """Form validation."""
        with transaction.atomic():
            response = super().form_valid(form)
            if self.object.is_posted and self.object.personal_account:
                self.object.personal_account.balance += self.object.amount
                self.object.personal_account.save()
            return response


class CashBoxExpenseCreateView(LoginRequiredMixin, RoleRequiredMixin, CreateView):
    """CashBox Expense Create View."""

    model = CashBox
    form_class = CashBoxExpenseForm
    template_name = "core/adminlte/cashbox_expense_form.html"
    success_url = reverse_lazy("finance:cashbox_list")
    permission_required = "has_cashbox"

    def get_initial(self):
        """Get initial data for the template."""
        initial = super().get_initial()
        last = CashBox.objects.aggregate(Max("id"))["id__max"] or 0
        initial["number"] = f"{last + 1}".zfill(10)
        initial["date"] = timezone.now().date()
        initial["manager"] = self.request.user
        return initial

    def form_valid(self, form):
        """Form validation."""
        return super().form_valid(form)


class CashBoxUpdateView(LoginRequiredMixin, RoleRequiredMixin, UpdateView):
    """Depending on the type of article, we substitute the appropriate template."""

    model = CashBox
    permission_required = "has_cashbox"
    success_url = reverse_lazy("finance:cashbox_list")

    def get_form_class(self):
        """Get form class."""
        if self.object.article.type == "income":
            return CashBoxIncomeForm
        return CashBoxExpenseForm

    def get_template_names(self):
        """Get template names."""
        if self.object.article.type == "income":
            return ["core/adminlte/cashbox_income_form.html"]
        return ["core/adminlte/cashbox_expense_form.html"]


class CashBoxDetailView(LoginRequiredMixin, RoleRequiredMixin, DetailView):
    """Display detailed information about a cashbox transaction."""

    model = CashBox
    template_name = "core/adminlte/cashbox_detail.html"
    permission_required = "has_cashbox"
    context_object_name = "cashbox"


class ExportCashBoxExcelView(LoginRequiredMixin, RoleRequiredMixin, View):
    """Export filtered CashBox transactions to Excel."""

    permission_required = "has_cashbox"

    def get(self, request, *args, **kwargs):
        """Get."""
        queryset = self.filter_queryset(request)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Касса"  # noqa: RUF001

        headers = [
            "№",
            "Дата",
            "Статус",
            "Тип платежа",
            "Владелец",
            "Лицевой счет",
            "Приход/Расход",
            "Сумма (грн)",
            "Комментарий",
        ]
        sheet.append(headers)

        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for item in queryset:
            owner_name = "-"
            if (
                item.personal_account
                and item.personal_account.apartment
                and item.personal_account.apartment.owner
            ):
                owner_name = item.personal_account.apartment.owner.get_full_name()

            status = "Проведен" if item.is_posted else "Не проведен"  # noqa: RUF001

            type_display = (
                "Приход"
                if item.article.type == Article.ArticleType.INCOME
                else "Расход"
            )

            amount = item.amount
            if item.article.type == Article.ArticleType.EXPENSE:
                amount = -amount

            row = [
                item.number,
                item.date.strftime("%d.%m.%Y"),
                status,
                item.article.name,
                owner_name,
                item.personal_account.number if item.personal_account else "-",
                type_display,
                amount,
                item.comment,
            ]
            sheet.append(row)

        column_widths = [15, 12, 12, 25, 30, 20, 15, 15, 40]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="cashbox_export.xlsx"'

        workbook.save(response)
        return response

    def filter_queryset(self, request):
        """Replicate the filtering logic from Datatables."""
        queryset = CashBox.objects.select_related(
            "article", "personal_account", "personal_account__apartment__owner"
        ).order_by("-date", "-id")

        number = request.GET.get("number", "").strip()
        date = request.GET.get("date", "").strip()
        is_posted = request.GET.get("is_posted", "").strip()
        article = request.GET.get("article", "").strip()
        owner = request.GET.get("owner", "").strip()
        account = request.GET.get("personal_account", "").strip()
        type_op = request.GET.get("type", "").strip()

        q = Q()

        if number:
            q &= Q(number__icontains=number)

        try:
            import re

            if re.match(r"^\d{1,2}\.\d{1,2}\.\d{4}$", date):
                day, month, year = map(int, date.split("."))
                filter_date = datetime.date(year, month, day)
                q &= Q(date=filter_date)
        except (ValueError, AttributeError):
            pass

        if is_posted:
            q &= Q(is_posted=(is_posted.lower() == "true"))

        if article:
            q &= Q(article_id=article)

        if owner:
            q &= Q(personal_account__apartment__owner_id=owner)

        if account:
            q &= Q(personal_account__number__icontains=account)

        if type_op:
            q &= Q(article__type=type_op)

        return queryset.filter(q)
